# SweepWidthParser.py
# Written by Yaron Milwid (2016), edited by Steven Edwards (2022)

import re
import tkinter
from tkinter import filedialog
import os
from os import listdir, makedirs
from os.path import isfile, exists, isdir
import numpy as np
from scipy.optimize import curve_fit
from scipy.integrate import quad
import matplotlib.pyplot as plt
from openpyxl import *
from weakref import proxy

class plant:
	def __init__ (self,  x,  y, field,type_a, type_overall):
		self.field = field#field containing artifact
		self.x = x#horizontal location
		self.y = y#location along transect
		self.d=(0.2*float(abs(x)))#amount of underestimation we allow
		self.found_count = 0#number of times this artifact has been found
		self.found_in_this_row =0#flag to ensure artifact only found once each traversal
		self.t = type_a#artifact type
		self.overall = type_overall#super-group for artifact
	def check_if_correct(self,x,this_row,walk_count):
		'''
			inputs: x=float, location along x axis,
				this_row = list, flag to determine whether to increment walk_count
				walk_count= dict containing walk counts for each field, only increment
					if this is the first time an artifact is found for this traversal
			outputs: bool: True=>found, False=>not found
			checks whether the object at the given (x,y) location is within the acceptable bounds to be
				the current plant, if so increments its found count and returns true, else returns false
		'''
		x = abs(x) #we don't care about the direction since the transect can be traversed
                            #in either direction
		xv=-300
		if x>0:
			xv = max(np.round(x),1)#round to nearest int
		elif x == 0:
			xv = 0
		sx=max(np.round(abs(self.x)),1)#round location of artifact to nearest int, and make positive
		dist = abs(sx)-abs(xv)#find offset from artifact
		if ((dist <=self.d and abs(sx)>=abs(xv)) or (abs(sx)==abs(xv))):#only allow underestimation
                    #or equivalence
			if self.found_in_this_row == 0:
				if this_row[0] ==0:
					walk_count[self.field]+=1
					this_row[0]=1
				self.found_count += 1
				self.found_in_this_row = 1
				return True
			else:
				return False
		else:
			return False

def get_dirs(dirs):
	''' input: dirs = empty list
		output: dirname= name of directory containing subdirectories with data 
		files and master files
	
		function gets user to select a folder, and populates dirs list with the
		names of all the subdirectories in the directory
	'''
	root = tkinter.Tk()
	dirname = filedialog.askdirectory(parent=root,initialdir="/",title='Please select a directory')
	root.destroy()
	
	for f in listdir(dirname):
		if isdir(dirname+"/"+f):
			dirs.append(dirname+"/"+f)
	
	return dirname

def get_files (dirname, logs):
	''' input: logs= empty list, 
			   dirname= name of directory containing files
		output: master_file = file containing the actual locations of artifacts
		
		Takes in a directory name and populates the logs list with the experiment
		data files as well as returns the file containing the mapping of the 
		artifacts
	'''
	master = 'master'
	field_data = 'Calibration\s+Records'
	
	master_file = ''
	
	for f in listdir(dirname):
		if isfile(dirname+"/"+f):
			if re.search(master, f, re.IGNORECASE) and re.search('csv', f,re.IGNORECASE):
				master_file = dirname+"/"+f
			elif re.search(field_data,f,re.IGNORECASE) and re.search('xls',f,re.IGNORECASE):
				logs.append(dirname+"/"+f)
	return master_file

def read_master(master_file, artifact_count, to_be_found, walk_count, fields):
	''' input: master_file = str with the file path of the known artifacts, 
		artifact_count =empty dict, will be indexed by artifact type 
		and field containing number of each type of artifact in that field, 
		to_be_found = empty dict, will contain all artifacts
		walk_count = empty dict, will be indexed by field
		fields=dict indexed by fields that will contain number of segments in
		each field
		output: none
		initializes the walk_count list, traverses the master_file csv and populates 
		input dictionaries
	'''

	mf = open(master_file, 'rbU')
	nf = 0
	ent = '(.+),(\d+),(\d+),(\S+),(\d+),(.+),(.+),(\d+)'
	art_pat = '([\D\s]+)(\d*)'
	for entry in mf:
		field = ''
		y= -500
		x = -500
		art = ''
		entry_data = re.search(ent, entry)
		#get individual elements from entry in master file
		if entry_data:
			overall = str(entry_data.group(7)).lower()
			num_seg = entry_data.group(8)
			field = str(entry_data.group(6))
			seg= float(entry_data.group(2))
			y = float(entry_data.group(3))
			x = float(entry_data.group(5))
			if entry_data.group(4) == 'l' or entry_data.group(4) == 'L':
				x = -x
			art = re.search(art_pat,entry_data.group(1)).group(1).lower().strip()
			if field not in fields:#populate the fields and walk_count dictionaries
				if not field.isdigit():
					field=field.lower()
				walk_count[field]=0
				nf +=1
				fields[field]=num_seg

			if overall in artifact_count:#update artifact_count dictionary
				if field in artifact_count[overall]:
					artifact_count[overall][field] +=1
				else:
					artifact_count[overall][field]=1
			else:
			   artifact_count[overall]={}
			   artifact_count[overall][field]=1
			#create plant object and add it to dictionary
			temp_plant = plant(x,y,field,art,overall)
			if art not in to_be_found:
				to_be_found[art]={}
			if field not in to_be_found[art]:
				to_be_found[art][field]={}
			if seg not in to_be_found[art][field]:
				to_be_found[art][field][seg]={}
			if int(y) not in to_be_found[art][field][seg]:
				to_be_found[art][field][seg][int(y)]=[]
			to_be_found[art][field][seg][int(y)].append(temp_plant)
	mf.close()

def read_field_data(f,walk_count, to_be_found, fields):
	'''
	inputs: f is the file containing data dumped from ipad
		walk_count = dict that keeps track of the number of times each
		#field has been walked, 
		to_be_found= dictionary with the plant objects
		fields= list of fields
	#outputs:none
		parses the walking data and compares against the artifacts in to_be_found
	'''
	f_number= -300
	pats = {'trans':r'TDist\s+(\d+)','seg':r'Segment\s+(\d+)','art':r'Artifact\s+(\d+)','direct':r'Direction\s+(\d+)','lat':r'Distance\s+(\d+)','field':r'Ground\s+Cover','tlen':r'Transect\s+Length'}

	cols={}
	cols['trans']={}
	cols['seg']={}
	cols['art']={}
	cols['direct']={}
	cols['lat']={}
	cols['field']=' '
	cols['tlen']=' '
	wb = load_workbook(filename=f, read_only = True)
	ws=wb.active
	for row in ws.rows:#map columns to artifact number, property pairs
		for cell in row:
			for p_k in pats:
				c_num = re.search(pats[p_k], cell.value)
				if c_num:
					if p_k == 'field':#only 1 entry per row
						cols[p_k]=cell.column
						continue 
					elif p_k == 'tlen':#only 1 entry per row
						cols[p_k]=cell.column
						continue
					cols[p_k][int(c_num.group(1))]=cell.column
					continue
		break #only want to look at the first row for header row
	max_num_arts = max(cols['art'].keys())
	ws.calculate_dimension(force = True)
	for r in xrange(ws.min_row+1, ws.max_row+1):
		this_row =[0] #use list for mutability
		found_list = []#use to keep track of which artifacts have been found
		f_number = str(ws.cell(row=r, column= cols['field']).value)
		if (not f_number.isdigit()):
			f_number = f_number.lower()#use for ease of matching
		for i in fields.keys():
			if re.search(i,f_number) != None:#official field name must
				#be smaller than surveyors' description, or the same as
				#surveyors' decriptions, and there must be a high level 
				#of confidence that the description will show up in every
				#description of htat field and no other field
				f_number= i
				break

		for i in xrange(1,max_num_arts+1):#traverse the row checking each artifact
			found_flag = 0
			if (((ws.cell(row = r, column = cols['art'][i]).value != None) and 
				((i in cols['trans']) and ws.cell(row=r, column=cols['trans'][i]).value != None) 
				and ((i in cols['seg']) and ws.cell(row=r, column=cols['seg'][i]).value)!=None)
				and ((i in cols['lat']) and ws.cell(row=r, column=cols['lat'][i]).value != None) and 
				((i in cols['direct']) and ws.cell(row=r, column= cols['direct'][i]).value != None)):
				#we require all data for a particular sighting except direction in order to be considered valid

				art = ws.cell(row = r, column = cols['art'][i]).value.lower()
				x = ws.cell(row=r, column=cols['lat'][i]).value
				seg_len=int(ws.cell(row=r, column=cols['tlen']).value)/int(fields[f_number])

				#surveyor's can traverse transect in either direction and we do not know how they will record
				#distance, so be lenient
				y = int(ws.cell(row=r, column=cols['trans'][i]).value)
				s = ws.cell(row=r, column=cols['seg'][i]).value
				s_op = int(fields[f_number])-int(s)+1
				y_op=seg_len-int(y)+1
				

				for a in to_be_found.keys():#search for artifact in the overall description
					#overall description should be as verbose as possible
					if re.search(art,a):
						if f_number in to_be_found[a].keys():
							if found_flag ==0 and s in to_be_found[a][f_number].keys() and y in to_be_found[a][f_number][s].keys():
								for p in to_be_found[a][f_number][s][y]:
									if p.check_if_correct(x,this_row,walk_count):
										found_list.append(proxy(p))
										found_flag =1
										break
							if found_flag==0 and s_op in to_be_found[a][f_number].keys() and y in to_be_found[a][f_number][s_op].keys():
								for p in to_be_found[a][f_number][s_op][y]:
									if p.check_if_correct(x,this_row,walk_count):
										found_list.append(proxy(p))
										found_flag =1
										break
							if found_flag==0 and s in to_be_found[a][f_number].keys() and y_op in to_be_found[a][f_number][s].keys():
								for p in to_be_found[a][f_number][s][y_op]:
									if p.check_if_correct(x,this_row,walk_count):
										found_list.append(proxy(p))
										found_flag =1
										break
							if found_flag==0 and s_op in to_be_found[a][f_number].keys() and y_op in to_be_found[a][f_number][s_op].keys():
								for p in to_be_found[a][f_number][s_op][y_op]:
									if p.check_if_correct(x,this_row,walk_count):
										found_list.append(proxy(p))
										found_flag =1
										break
							if found_flag ==1:
								break
							
		for i in found_list:
			i.found_in_this_row = 0#reset found_in_this_row flags so that can find artifacts in next row

def aggregate(walk_count, to_be_found, possible, found, dirname):
	'''
	inputs:walk_count=dict (the number of times each field was walked
		to_be_found=dict(the dictionary containing all of the plants to be found), 
		possible=dictionary to map each artifact-field-x location to the number
		of times it could be found
		found = dictionary to map each artifact-field-x location to the number
		of times it was found
		dirname = directory we are currently looking at
	
	outputs:none
	
	Stores the found information and walk count for this directory in a file
	Aggregates teh foudn and possible information with other directories
	'''
	if not exists(dirname+"/output"):
		makedirs(dirname+"/output")
	#store data in to_be_found and walk_count
	f=dirname+"/output/found_summary.csv"
	mf=open(f,'a')
	mf.write("article,field,segment,y,x,found_count\n")
	for a in to_be_found.keys():
		for f in to_be_found[a].keys():
			for s in to_be_found[a][f].keys():
				for y in to_be_found[a][f][s].keys():
					for p in to_be_found[a][f][s][y]:
						mf.write(p.t+","+p.field+","+str(s)+","+str(p.y)+","+str(p.x)+","+str(p.found_count)+"\n")
	mf.write("###############,############,#############,############\n")
	mf.write("field, walk count\n")
	for f in walk_count:
		mf.write(f+","+str(walk_count[f])+"\n")
	mf.close()
	
	for art in to_be_found:
		#aggregate all data for each artifact, field, location combination
		for f in to_be_found[art]:
			for s in to_be_found[art][f]:
				for y in to_be_found[art][f][s]:
					for p in to_be_found[art][f][s][y]:
						if p.overall not in possible:
							possible[p.overall]={}
							found[p.overall]={}
							for k in walk_count.keys():
								possible[p.overall][k]={}
								found[p.overall][k]={}
						if p.x not in possible[p.overall]:
							possible[p.overall][f][p.x]=0
							found[p.overall][f][p.x]=0
		#if a field was walked, any artifact in it could be found on each traversal
						possible[p.overall][f][p.x]+=walk_count[f]
						found[p.overall][f][p.x]+=p.found_count

def calculate_probability(probability_f_a, probability_a,possible, found,dirname):
	'''
	inputs: 
		probability_f_a=empty dict(will contain the probability values for each
		artifact separated by field and artifact type), 
		probability_a=empty dict(will contain the probability values for
		each artifact separated by the artifact type)
		possible = dictionary containing the number of times each artifact-field-x
		combination could be found
		found = dictionary containing the number of times each artifact-field-x
		combination was found
		dirname = name of directory containing subdirectories
	outputs: none
	calculates the probability of an article being found. Stores found statistics data 
		for each artifact as well as probabilities of being found
	'''
	if not exists(dirname+"/output"):
		makedirs(dirname+"/output")
	
	
	for art in possible.keys():#calculate the probabilities by field and artifact
		probability_f_a[art]={}
		x_poss = {}
		x_found = {}
		for f in possible[art].keys():
			probability_f_a[art][f]={}
			for x in possible[art][f].keys():
				if possible[art][f][x] !=0:
					probability_f_a[art][f][float(x)]=(float(found[art][f][x])/possible[art][f][x])
					if x not in x_poss:
						x_poss[x]=0
						x_found[x]=0
					x_poss[x] +=possible[art][f][x]
					x_found[x]+=found[art][f][x]
		for x in x_poss.keys():
			if art not in probability_a.keys():
				probability_a[art]={}
			probability_a[art][float(x)]=float(x_found[x])/x_poss[x]
		#store probabilities
	mf=open(dirname+"/output/probabilities.csv",'w')
	mf.write("artifact,x,probability\n")
	for art in probability_a.keys():
		for x in probability_a[art].keys():
			mf.write(art+","+str(x)+","+str(probability_a[art][x])+"\n")
	mf.write("\n\n\nart,f,x,probability\n")
	for art in probability_f_a.keys():
		for f in probability_f_a[art].keys():
			for x in probability_f_a[art][f].keys():
				mf.write(art+","+str(f)+","+str(x)+","+str(probability_f_a[art][f][x])+"\n")
	mf.close()

def curve_to_fit(x, b, k):
	'''
	probability of detection equation
	'''
	return b*np.exp(-k*np.power(x,2))

def fit(dirname, probability_f_a, probability_a, w_a):
	'''
	Fit a curve to the probabilities for both artifacts
	and artifact-field combinations. Calculate the sweep
	width based on this curve, plot the data and the
	curve and store in width.csv
	'''
	d = dirname+"/output"
	if not exists(d):
		  makedirs(d)
	wf = open(dirname+"/output/"+"width.csv", 'w')
	wf.write("Artifact Type, Field #, Width, b,b_cov, k,k_cov\n")
	for art in probability_a:#fit a curve and plot data, combining all fields
		range_x=probability_a[art].keys()
		range_x.sort()
		pvals=[]
		for x in range_x:#need to do this as a separate step to ensure ordering
			pvals.append(probability_a[art][x])
		popt,pcov=curve_fit(curve_to_fit,range_x,pvals,p0=[0.63,0.01], bounds= (0,[1,1]))
		b= popt[0]
		k=popt[1]
		
		width,abserr = quad(curve_to_fit,-100,100,args=(b,k))#numeric integration to find width
		x= np.linspace(-20,20)
		wf.write(art+",all,"+str(width)+","+str(b)+","+str(pcov[0])+","+str(k)+","+str(pcov[1])+"\n")
		plt.plot(range_x,pvals, 'ro')#plot data
		plt.plot(x,curve_to_fit(x,*popt),'g',lw=2)#plot fitted curve
		plt.ylabel('P(r)')
		plt.xlabel('range(m)')
		plt.title("P(r) for all fields and artifact = "+art)
		name=d+"/"+art+"_all_fields.png"
		plt.savefig(name)
		plt.close()
		w_a[art]=width
		for fnum in probability_f_a[art]:#separate plots by field
			range_x = probability_f_a[art][fnum].keys()
			range_x.sort()
			pvals=[]
			for x in range_x:
				pvals.append(probability_f_a[art][fnum][x])
			if len(range_x)>0:
					popt=[]
					pcov=[]
					popt,pcov=curve_fit(curve_to_fit,range_x,pvals,p0=[0.63,0.01], bounds= (0,[1,1]))
					b= popt[0]
					k= popt[1]
					x= np.linspace(-20,20)
					width,abserr = quad(curve_to_fit,-100,100,args=(b,k))
					wf.write(art+","+str(fnum)+","+str(width)+","+str(b)+","+str(np.sqrt(np.diag(pcov,0)))+","+str(k)+","+str(np.sqrt(np.diag(pcov,1)))+"\n")
					plt.plot(range_x,pvals, 'ro')
					plt.plot(x,curve_to_fit(x,*popt),'g',lw=2)
					plt.ylabel('P(r)')
					plt.xlabel('range(m)')
					plt.title("P(r) for field = "+str(fnum) +" and artifact = "+art)
					name=str(d)+"/"+str(art)+"_field#"+str(fnum)+".png"
					plt.savefig(name)
					plt.close()
	wf.close()

def get_weighted_average(dirname,w_a,artifact_count):
	'''
	Calculate the average width, weighted by number of
	artifacts of that type and store in the width file
	'''

	total_count = 0
	width_num = 0
	for art in w_a:
		total_art =0
		for f in artifact_count[art].keys():
			total_count += artifact_count[art][f]
			total_art +=artifact_count[art][f]
		width_num += (total_art * w_a[art])
	weighted_average = float(width_num)/total_count
	f=dirname+"/output/width.csv"
	ff=open(f,'a')
	ff.write("weighted_average,all,")
	ff.write(str(weighted_average))
	ff.write("\n")
	ff.close()

probability_f_a = {}
probability_a = {}
w_a={}
possible = {}
found = {}
dirs = []
artifact_count = {}

dirname = get_dirs(dirs)

for d in dirs:
	logs = []
	walk_count = {}
	to_be_found = {}
	fields = {}
	
	mf = "C:/Users/W0304792/Desktop/sweep2/Experiment 1/master.xlsx"
	read_master(mf,artifact_count,to_be_found, walk_count,fields)
	for l in logs:#read multiple logs
		read_field_data(l,walk_count,to_be_found,fields)
	aggregate(walk_count, to_be_found, possible, found,d)

calculate_probability(probability_f_a,probability_a, possible, found, dirname)
fit(dirname, probability_f_a, probability_a, w_a)
get_weighted_average(dirname,w_a,artifact_count)